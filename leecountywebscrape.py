import openpyxl

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException

import time

# these options allow the web page to stay open #
options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option('detach', True)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# we need to open the page first and find the elements before looping through #
url = "https://www.example.com"
driver.get(url)
time.sleep(1)
street_input = driver.find_element(By.ID, "ctl00_BodyContentPlaceHolder_WebTab1_tmpl0_AddressTextBox")      
  
wrkbk = openpyxl.load_workbook("NameOfSpreadsheet.xlsx")
  
sheet = wrkbk.active

row = sheet.max_row
column = sheet.max_column

# in my case, range(2, 2) is the indicated row and column to start at #
for i in range(2, 2): 
    # open the page again to find the elements because it will result in 'not found' or 'time out' errors otherwise #
    driver.get(url)
    
    street_input = driver.find_element(By.ID, "ctl00_BodyContentPlaceHolder_WebTab1_tmpl0_AddressTextBox")
    
    time.sleep(1)

    cell = sheet.cell(row = i, column = 18) 
    street_input.send_keys(cell.value) # send_keys enters the value of the spreadsheet cell into the form input #
    street_input.send_keys(Keys.ENTER) # this send_keys hits the keyboard ENTER key for me to initiate the search #
    
    # there should be some sort of exception catching here for when no results are found #
    results = WebDriverWait(driver, 15).until(ec.presence_of_element_located((By.XPATH, "/html/body/form/div[5]/div/div/div[1]/div/div/div/div[1]/div[1]/table/tbody/tr/td[1]/div/div[2]"))).text
    
    # appending the info I needed to a url. I already knew the url, I just needed an Id from the page I was searching for # 
    base_url = "https://www.example.com?ID="
    full_url = f"{base_url}{results}"
    print(full_url)


    # had this in place before url appendage. used for navigating to the subsequent details page and printing the whole url - probs unnecessary since I already knew the base url #
    #results = driver.find_element(By.XPATH, "/html/body/form/div[5]/div/div/div[1]/div/div/div/div[1]/div[1]/table/tbody/tr/td[4]/div/div[1]/a").click()
    #driver.switch_to.window(driver.window_handles[-1])
    #print(driver.current_url)


